import pandas as pd
import io

from openpyxl.styles import Font, PatternFill
from openpyxl import Workbook

def build_excel(sku_details):
    # Initialize workbook and select active worksheet
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Product Images QA"

    # Define headers matching your HTML table
    headers = [
        "SKU", "SKU Name", "Orientation Count", 
        "Product Color Count", "Document Type Count",
        "Document Type Breakdown",
        "Master Object Name Count", "Type Count", 
        "Total Images", "Notes"
    ]

    # Add headers to the Excel sheet
    sheet.append(headers)

    # Define header styles (bold, background color)
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="0096D6")
    for col in range(1, len(headers) + 1):
        sheet.cell(row=1, column=col).font = header_font
        sheet.cell(row=1, column=col).fill = header_fill

    # Add SKU data rows
    for sku in sku_details:
        counts = sku["counts"]
        document_type_breakdown = "; ".join([f"{k}: {v}" for k, v in counts["documentTypeDetail_breakdown"].items()])
        
        row = [
            sku["sku"], sku["sku_name"], 
            counts["orientation_count"], counts["productColor_count"], 
            counts["documentTypeDetail_count"],
            document_type_breakdown,
            counts["masterObjectName_count"], counts["type_count"], 
            sku["image_count"], sku.get("notes", "")
        ]
        sheet.append(row)

    # Apply conditional formatting (coloring zero values red)
    red_font = Font(color="FF0000")
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=3, max_col=9):
        for cell in row:
            if cell.value == 0:
                cell.font = red_font

    # Save the workbook to an in-memory buffer
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
